import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook

class Workbook:
    def __init__(self, file_path:Path|str):
        """
        :param file_path: Path to the Excel file (.xls, .xlsx, .xlsm, etc.)
        """

        if isinstance(file_path, Path):
            self.file_path = file_path
        else:
            self.file_path = Path(file_path)

    def sheet_names(self) -> list[int|str]:
        """
        Returns a list of sheet names from an Excel file.
        :return: List of sheet names or None if error occurs
        """
        try:
            # Validate file existence
            if not os.path.isfile(self.file_path):
                raise FileNotFoundError(f"File not found: {self.file_path}")
            sheet_names: list[int|str]
            # Load Excel file
            with pd.ExcelFile(self.file_path, 'openpyxl') as excel_file:
                sheet_names =  excel_file.sheet_names
            # Return list of sheet names
            return sheet_names

        except FileNotFoundError as fnf_err:
            print(f"Error: {fnf_err}")
        except ValueError as val_err:
            print(f"Invalid file format: {val_err}")
        except Exception as e:
            print(f"Unexpected error: {e}")

        return None
    
    def xlsm_sheet_names(self):
        try:
            # Load the workbook (read-only mode for efficiency)
            workbook = load_workbook(filename=self.file_path, read_only=True, keep_vba=True)

            # Get all sheet names
            sheet_names = workbook.sheetnames

            return sheet_names
        except FileNotFoundError:
            print(f"Error: File '{self.file_path}' not found.")
        except Exception as e:
            print(f"Error reading file: {e}")

